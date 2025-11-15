"""
Gerador de Relatório Excel MAPA.
Formato oficial do Ministério da Agricultura.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict
from datetime import datetime


class MAPAReportGenerator:
    """
    Gera relatório Excel no formato oficial MAPA.
    """

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Relatório MAPA"

    def generate(
        self,
        user_info: dict,
        period: str,
        rows: List[Dict],
        output_path: str
    ) -> str:
        """
        Gera relatório Excel completo.

        Args:
            user_info: Dados do usuário/empresa (nome, CNPJ, email, etc.)
            period: Período do relatório (ex: "Q1-2025")
            rows: Lista de dicionários com dados agregados
            output_path: Caminho do arquivo de saída

        Returns:
            Caminho do arquivo gerado
        """
        # Header do relatório
        self._write_header(user_info, period)

        # Cabeçalhos das colunas
        self._write_column_headers()

        # Dados
        self._write_data_rows(rows)

        # Formatação
        self._apply_formatting()

        # Salvar
        self.wb.save(output_path)
        return output_path

    def _write_header(self, user_info: dict, period: str):
        """Escreve header do relatório (linhas 1-12)"""

        # Linha 1: Título
        self.ws['A1'] = "MINISTÉRIO DA AGRICULTURA, PECUÁRIA E ABASTECIMENTO"
        self.ws['A1'].font = Font(bold=True, size=14)
        self.ws['A1'].alignment = Alignment(horizontal='center')

        # Linha 2: Subtítulo
        self.ws['A2'] = "CFIC - CONTROLE FISCAL DE INSUMOS AGRÍCOLAS"
        self.ws['A2'].font = Font(bold=True, size=12)
        self.ws['A2'].alignment = Alignment(horizontal='center')

        # Linha 3: Tipo de relatório
        self.ws['A3'] = "RELATÓRIO TRIMESTRAL DE PRODUÇÃO E COMERCIALIZAÇÃO"
        self.ws['A3'].font = Font(bold=True, size=11)
        self.ws['A3'].alignment = Alignment(horizontal='center')

        # Linha 5: Informações da empresa
        self.ws['A5'] = f"Estabelecimento: {user_info.get('company_name', '')}"
        self.ws['A6'] = f"CNPJ: {user_info.get('cnpj', 'N/A')}"
        self.ws['A7'] = f"Email: {user_info.get('email', '')}"

        # Linha 8: Período
        quarter, year = period.split('-')
        self.ws['A8'] = f"Trimestre: {quarter}"
        self.ws['B8'] = f"Ano: {year}"

        # Linha 10: Data de geração
        self.ws['A10'] = f"Data de Geração: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    def _write_column_headers(self):
        """Escreve cabeçalhos das colunas (linha 14)"""

        headers = [
            "Material/Produto",
            "Nº Registro MAPA",
            "Referência",
            "Unidade",
            "Compra Nacional (Ton)",
            "Compra Importada (Ton)",
            "Total (Ton)",
            "NF-es Origem"
        ]

        row = 14
        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    def _write_data_rows(self, rows: List[Dict]):
        """Escreve linhas de dados (a partir da linha 15)"""

        start_row = 15

        for idx, row_data in enumerate(rows):
            row_num = start_row + idx

            # Dados da linha
            mapa_reg = row_data.get("mapa_registration", "")
            reference = row_data.get("product_reference", "")
            unit = row_data.get("unit", "Tonelada")
            qty_domestic = row_data.get("quantity_domestic", "0")
            qty_import = row_data.get("quantity_import", "0")

            # Calcular total
            try:
                total = float(qty_domestic) + float(qty_import)
            except:
                total = 0.0

            # NF-es origem (lista -> string)
            nfes = row_data.get("source_nfes", [])
            nfes_str = ", ".join(nfes) if nfes else ""

            # Escrever células
            self.ws.cell(row=row_num, column=1).value = mapa_reg  # Material/Produto
            self.ws.cell(row=row_num, column=2).value = mapa_reg  # Nº Registro
            self.ws.cell(row=row_num, column=3).value = reference  # Referência
            self.ws.cell(row=row_num, column=4).value = unit  # Unidade
            self.ws.cell(row=row_num, column=5).value = float(qty_domestic)  # Nacional
            self.ws.cell(row=row_num, column=6).value = float(qty_import)  # Importado
            self.ws.cell(row=row_num, column=7).value = total  # Total
            self.ws.cell(row=row_num, column=8).value = nfes_str  # NF-es

            # Formatar números
            for col in [5, 6, 7]:
                self.ws.cell(row=row_num, column=col).number_format = '0.000'

    def _apply_formatting(self):
        """Aplica formatação geral (bordas, larguras, etc.)"""

        # Bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplicar bordas na área de dados
        for row in self.ws.iter_rows(min_row=14, max_row=self.ws.max_row, min_col=1, max_col=8):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Ajustar larguras das colunas
        column_widths = {
            'A': 25,  # Material/Produto
            'B': 20,  # Nº Registro
            'C': 30,  # Referência
            'D': 12,  # Unidade
            'E': 18,  # Nacional
            'F': 18,  # Importado
            'G': 15,  # Total
            'H': 30   # NF-es
        }

        for col_letter, width in column_widths.items():
            self.ws.column_dimensions[col_letter].width = width


def generate_mapa_report(
    user_info: dict,
    period: str,
    rows: List[Dict],
    output_path: str
) -> str:
    """
    Função helper para gerar relatório MAPA.

    Args:
        user_info: Informações do usuário
        period: Período (ex: "Q1-2025")
        rows: Dados agregados
        output_path: Caminho do arquivo de saída

    Returns:
        Caminho do arquivo gerado
    """
    generator = MAPAReportGenerator()
    return generator.generate(user_info, period, rows, output_path)
