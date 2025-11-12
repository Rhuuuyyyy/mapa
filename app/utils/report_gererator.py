import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any
import os
from datetime import datetime


class MAPAReportGenerator:
    """
    Gerador de Relatório Trimestral MAPA
    Formato: Relatório de Produção, Importação e Exportação
    """
    
    def __init__(self, user, period: str):
        self.user = user
        self.period = period  # Ex: Q1-2025
        self.nfe_data_list = []
        self.report_dir = "reports"
        os.makedirs(self.report_dir, exist_ok=True)
    
    def add_nfe_data(self, nfe_data: Dict[str, Any]):
        """Adiciona dados de uma NF-e ao relatório"""
        self.nfe_data_list.append(nfe_data)
    
    def _extract_guarantees_from_description(self, description: str) -> Dict[str, str]:
        """
        Extrai as garantias (nutrientes) da descrição do produto
        Ex: "N03 P04 K15 + 12Ca MG" -> N=03, P2O5=04, K2O=15, Ca=12
        """
        import re
        
        guarantees = {
            'N': '',
            'P2O5': '',
            'K2O': '',
            'Ca': '',
            'Mg': '',
            'S': '',
            'B': '',
            'Cl': '',
            'Cu': '',
            'Fe': '',
            'Mn': '',
            'Mo': '',
            'Ni': '',
            'Se': '',
            'Si': '',
            'Zn': ''
        }
        
        # Padrões de extração
        patterns = {
            'N': r'N\s*(\d+)',
            'P2O5': r'P\s*(\d+)',
            'K2O': r'K\s*(\d+)',
            'Ca': r'(\d+)\s*Ca',
            'Mg': r'Mg\s*(\d+)|(\d+)\s*Mg',
        }
        
        for nutrient, pattern in patterns.items():
            match = re.search(pattern, description, re.IGNORECASE)
            if match:
                value = match.group(1) if match.group(1) else match.group(2) if len(match.groups()) > 1 else ''
                if value:
                    guarantees[nutrient] = value
        
        return guarantees
    
    def _parse_products_for_mapa(self) -> List[Dict[str, Any]]:
        """Converte dados das NF-es para o formato do relatório MAPA"""
        mapa_rows = []
        
        for nfe in self.nfe_data_list:
            produtos = nfe.get('produtos', [])
            emitente = nfe.get('emitente', {})
            destinatario = nfe.get('destinatario', {})
            
            for produto in produtos:
                # Extrai garantias da descrição
                descricao = produto.get('descricao', '')
                guarantees = self._extract_guarantees_from_description(descricao)
                
                # Extrai registro MAPA da descrição
                import re
                registro_match = re.search(r'REG[\.:]?\s*MAPA[:\s]+([A-Z]{2}\s*\d+-\d+\.\d+-\d+)', descricao, re.IGNORECASE)
                registro_mapa = registro_match.group(1) if registro_match else ''
                
                # Converte quantidade para toneladas
                quantidade_str = produto.get('quantidade', '0').replace(',', '.')
                try:
                    quantidade = float(quantidade_str)
                except:
                    quantidade = 0.0
                
                row = {
                    # Identificação
                    'Tipo': 'Fertilizante',  # ou extrair do produto
                    'N_REGISTRO_PRODUTO': registro_mapa,
                    'AUTORIZACAO': '',
                    
                    # Garantias (Nutrientes) - em %
                    'N_Total': guarantees.get('N', ''),
                    'N_Solub': '',
                    'P2O5': guarantees.get('P2O5', ''),
                    'K2O': guarantees.get('K2O', ''),
                    'Ca': guarantees.get('Ca', ''),
                    'Mg': guarantees.get('Mg', ''),
                    'S': guarantees.get('S', ''),
                    'B': guarantees.get('B', ''),
                    'Cl': guarantees.get('Cl', ''),
                    'Cu': guarantees.get('Cu', ''),
                    'Fe': guarantees.get('Fe', ''),
                    'Mn': guarantees.get('Mn', ''),
                    'Mo': guarantees.get('Mo', ''),
                    'Ni': guarantees.get('Ni', ''),
                    'Se': guarantees.get('Se', ''),
                    'Si': guarantees.get('Si', ''),
                    'Zn': guarantees.get('Zn', ''),
                    
                    # Quantidade Produzida/Importada
                    'Unidade': 'TON',
                    'Estoque_Inicial': '',
                    'Producao_EP_Adquirida_EC_Granel_Trimestre': quantidade if produto.get('unidade', '').upper() == 'TON' else '',
                    'Importacao_Quantidade': '',
                    'Importacao_Pais_Origem': '',
                    
                    # Quantidade Comercializada
                    'Outras_UFs_Quantidade': '',
                    'Outras_UFs_UFs': '',
                    'Exportacao_Quantidade': '',
                    'Exportacao_Pais_Destino': '',
                    'Na_UF_Quantidade': '',
                    
                    # Produto Devolvido
                    'Sem_Condicao_Revenda_Quantidade': '',
                    'Sem_Condicao_Revenda_Destinacao': '',
                    'Com_Condicao_Revenda_Quantidade': '',
                    'Com_Condicao_Revenda_Destinacao': '',
                    'Estoque_Final_Trimestre_Quantidade': '',
                    
                    # Metadados
                    'Descricao_Produto': descricao,
                    'Emitente_CNPJ': emitente.get('cnpj', ''),
                    'Emitente_Razao_Social': emitente.get('razao_social', ''),
                    'Destinatario_CNPJ': destinatario.get('cnpj_cpf', ''),
                    'Destinatario_Razao_Social': destinatario.get('razao_social', ''),
                    'Numero_NF': nfe.get('numero_nota', ''),
                    'Data_Emissao': nfe.get('data_emissao', ''),
                    'Chave_Acesso': nfe.get('chave_acesso', ''),
                }
                
                mapa_rows.append(row)
        
        return mapa_rows
    
    def generate_excel(self) -> str:
        """Gera o arquivo Excel no formato MAPA"""
        # Parse data
        mapa_data = self._parse_products_for_mapa()
        
        if not mapa_data:
            raise ValueError("Nenhum dado para gerar relatório")
        
        # Create DataFrame
        df = pd.DataFrame(mapa_data)
        
        # Create Excel file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Relatorio_MAPA_{self.period}_{timestamp}.xlsx"
        filepath = os.path.join(self.report_dir, f"user_{self.user.id}_{filename}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Relatório {self.period}"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title section
        ws.merge_cells('A1:Z1')
        title_cell = ws['A1']
        title_cell.value = "MINISTÉRIO DA AGRICULTURA, PECUÁRIA E ABASTECIMENTO – MAPA"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells('A2:Z2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = "COORDENAÇÃO DE FERTILIZANTES, INOCULANTES E CORRETIVOS - CFIC"
        subtitle_cell.font = Font(bold=True, size=12)
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells('A3:Z3')
        report_title_cell = ws['A3']
        report_title_cell.value = "Relatório Trimestral de Produção, Importação e Exportação"
        report_title_cell.font = Font(bold=True, size=12, color="0000FF")
        report_title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Company info section (row 5)
        ws['A5'] = "Estabelecimento:"
        ws['B5'] = self.user.company_name or self.user.full_name
        
        ws['P5'] = f"Registro nº:"
        ws['R5'] = f"UF:"
        
        ws['A6'] = "Endereço:"
        ws['P6'] = "CNPJ:"
        ws['B6'] = ""  # Can be filled from user data
        
        ws['A7'] = "E-Mail:"
        ws['B7'] = self.user.email
        ws['P7'] = "Fone:"
        ws['R7'] = f"Trimestre: {self.period}"
        ws['T7'] = "Ano:"
        
        # Merged info cells
        ws.merge_cells('B5:O5')
        ws.merge_cells('B6:O6')
        ws.merge_cells('B7:O7')
        
        # Main table headers (row 10)
        current_row = 10
        
        # Define column headers matching MAPA format
        headers = [
            'Tipo', 'Nº REGISTRO\nDE PRODUTO/\nAUTORIZAÇÃO',
            'N\nTotal', 'N\nSolúb', 'P₂O₅', 'K₂O', 'Ca', 'Mg', 'S', 'B', 'Cl',
            'Cu', 'Fe', 'Mn', 'Mo', 'Ni', 'Se', 'Si', 'Zn',
            'Unidade\nEstoque Inicial',
            'Produção EP /\nAdquirida EC\na granel no\nTrimestre\nQuantidade',
            'Importação do trimestre\nQuantidade', 'País de origem',
            'Outras UFs\nQuantidade', 'UFs',
            'Exportação\nQuantidade', 'País Destino',
            'Na UF\nQuantidade',
            'Sem condição de revenda\nQuantidade', 'Destinação',
            'Com condição de revenda\nQuantidade', 'Destinação',
            'Estoque Final no\nTrimestre\nQuantidade'
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows
        current_row += 1
        for row_data in mapa_data:
            col_idx = 1
            
            # Write each cell
            data_values = [
                row_data.get('Tipo', ''),
                row_data.get('N_REGISTRO_PRODUTO', ''),
                row_data.get('N_Total', ''),
                row_data.get('N_Solub', ''),
                row_data.get('P2O5', ''),
                row_data.get('K2O', ''),
                row_data.get('Ca', ''),
                row_data.get('Mg', ''),
                row_data.get('S', ''),
                row_data.get('B', ''),
                row_data.get('Cl', ''),
                row_data.get('Cu', ''),
                row_data.get('Fe', ''),
                row_data.get('Mn', ''),
                row_data.get('Mo', ''),
                row_data.get('Ni', ''),
                row_data.get('Se', ''),
                row_data.get('Si', ''),
                row_data.get('Zn', ''),
                row_data.get('Unidade', ''),
                row_data.get('Producao_EP_Adquirida_EC_Granel_Trimestre', ''),
                row_data.get('Importacao_Quantidade', ''),
                row_data.get('Importacao_Pais_Origem', ''),
                row_data.get('Outras_UFs_Quantidade', ''),
                row_data.get('Outras_UFs_UFs', ''),
                row_data.get('Exportacao_Quantidade', ''),
                row_data.get('Exportacao_Pais_Destino', ''),
                row_data.get('Na_UF_Quantidade', ''),
                row_data.get('Sem_Condicao_Revenda_Quantidade', ''),
                row_data.get('Sem_Condicao_Revenda_Destinacao', ''),
                row_data.get('Com_Condicao_Revenda_Quantidade', ''),
                row_data.get('Com_Condicao_Revenda_Destinacao', ''),
                row_data.get('Estoque_Final_Trimestre_Quantidade', ''),
            ]
            
            for value in data_values:
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                col_idx += 1
            
            current_row += 1
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save workbook
        wb.save(filepath)
        
        return filepath