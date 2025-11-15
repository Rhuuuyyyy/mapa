"""
MAPA Report Generator - Excel report generation in official MAPA format.

This module generates quarterly fertilizer reports in the exact format required by
Brazil's Ministry of Agriculture (MAPA) for raw materials.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Optional
from datetime import datetime
import os

from app.utils.mapa_mapper import MAPARow


class MAPAReportGenerator:
    """
    Generates Excel reports in official MAPA format.

    Creates quarterly reports matching the exact structure required by MAPA
    for fertilizer raw material producers, importers, and exporters.

    The report structure includes:
    - Header section (rows 1-4): Ministry title and report type
    - Company info (rows 7-12): Establishment details, contact, responsible person
    - Column headers (rows 14-16): Multi-row headers with 46 columns
    - Data rows (17+): Product entries with guarantees and quantities

    Usage:
        generator = MAPAReportGenerator(user, "3º-2025")
        generator.add_mapa_rows(rows)
        filepath = generator.generate_excel()
    """

    def __init__(self, user, period: str, company_info: Optional[dict] = None):
        """
        Initialize report generator.

        Args:
            user: User object with company information
            period: Report period (e.g., "1º-2025", "3º-2025")
            company_info: Optional dict with additional company data
                {
                    'registro_mapa': str,
                    'uf': str,
                    'endereco': str,
                    'telefone': str
                }
        """
        self.user = user
        self.period = period
        self.company_info = company_info or {}
        self.mapa_rows: List[MAPARow] = []
        self.report_dir = "reports"
        os.makedirs(self.report_dir, exist_ok=True)

    def add_mapa_rows(self, rows: List[MAPARow]) -> None:
        """
        Add MAPA rows to report.

        Args:
            rows: List of MAPARow objects to include in report
        """
        self.mapa_rows.extend(rows)

    def generate_excel(self) -> str:
        """
        Generate Excel file in official MAPA format.

        Returns:
            Path to generated Excel file

        Raises:
            ValueError: If no data to generate report
        """
        if not self.mapa_rows:
            raise ValueError("Nenhum dado para gerar relatório")

        wb = Workbook()
        ws = wb.active
        ws.title = "Fert. Min. (Mat.prima)"

        # Define styles
        self._define_styles()

        # Build report sections
        self._add_header_section(ws)
        self._add_company_info_section(ws)
        self._add_responsible_section(ws)
        self._add_column_headers(ws)
        self._add_data_rows(ws)

        # Adjust column widths
        self._adjust_column_widths(ws)

        # Save file
        filepath = self._save_workbook(wb)

        return filepath

    def _define_styles(self) -> None:
        """Define reusable cell styles"""
        # Title styles
        self.title_font = Font(bold=True, size=12, name='Arial')
        self.subtitle_font = Font(bold=True, size=11, name='Arial')
        self.header_font = Font(bold=True, size=9, name='Arial')

        # Data styles
        self.normal_font = Font(size=9, name='Arial')
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_alignment = Alignment(horizontal='left', vertical='center')

        # Borders
        self.thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Fills
        self.header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    def _add_header_section(self, ws) -> None:
        """Add report header (rows 1-5)"""
        # Row 1: Ministry title
        ws['A1'] = "MINISTÉRIO DA AGRICULTURA, PECUÁRIA E ABASTECIMENTO – MAPA"
        ws['A1'].font = Font(bold=True, size=14, name='Arial')
        ws['A1'].alignment = self.center_alignment

        # Row 2: Department
        ws['A2'] = "COORDENAÇÃO DE FERTILIZANTES, INOCULANTES E CORRETIVOS - CFIC"
        ws['A2'].font = Font(bold=True, size=12, name='Arial')
        ws['A2'].alignment = self.center_alignment

        # Row 4: Report type
        ws['A4'] = "Mapa Trimestral de Produção, Importação e Exportação"
        ws['A4'].font = Font(bold=True, size=12, name='Arial')
        ws['A4'].alignment = self.center_alignment

        # Row 5: Category
        ws['A5'] = "PARA ESTABELECIMENTOS PRODUTORES, IMPORTADORES E EXPORTADORES DE FERTILIZANTES MINERAIS ( Matéria-Prima )"
        ws['A5'].font = Font(bold=True, size=10, name='Arial')
        ws['A5'].alignment = self.center_alignment

    def _add_company_info_section(self, ws) -> None:
        """Add company information (rows 7-9)"""
        # Row 7: Establishment name and registration
        ws['A7'] = "Estabelecimento:"
        ws['A7'].font = self.subtitle_font

        # Company name (merged cells E7:Z7, not overlapping with other labels)
        company_name = self.user.company_name or self.user.full_name or ""
        ws.merge_cells('E7:Z7')
        ws['E7'] = company_name
        ws['E7'].font = self.normal_font

        # Registro label and value
        ws['AL7'] = "Registro nº:"
        ws['AL7'].font = self.subtitle_font

        registro = self.company_info.get('registro_mapa', '')
        ws.merge_cells('AM7:AR7')
        ws['AM7'] = registro
        ws['AM7'].font = self.normal_font

        # UF label and value
        ws['AS7'] = "UF:"
        ws['AS7'].font = self.subtitle_font

        uf = self.company_info.get('uf', '')
        ws['AT7'] = uf
        ws['AT7'].font = self.normal_font

        # Row 8: Address and CNPJ
        ws['A8'] = "Endereço:"
        ws['A8'].font = self.subtitle_font

        # Address (merged cells D8:AK8, not overlapping with other labels)
        endereco = self.company_info.get('endereco', '')
        ws.merge_cells('D8:AK8')
        ws['D8'] = endereco
        ws['D8'].font = self.normal_font

        # CNPJ label
        ws['AM8'] = "CNPJ:"
        ws['AM8'].font = self.subtitle_font

        # CNPJ value (merged cells)
        cnpj = getattr(self.user, 'cnpj', '') or ''
        ws.merge_cells('AO8:AT8')
        ws['AO8'] = cnpj
        ws['AO8'].font = self.normal_font

        # Row 9: Email, phone, trimester
        ws['A9'] = "E-Mail:"
        ws['A9'].font = self.subtitle_font

        # Email (merged cells C9:AC9, not overlapping)
        ws.merge_cells('C9:AC9')
        ws['C9'] = self.user.email or ""
        ws['C9'].font = self.normal_font

        # Phone label
        ws['AE9'] = "Fone:"
        ws['AE9'].font = self.subtitle_font

        # Phone value (merged cells)
        telefone = self.company_info.get('telefone', '')
        ws.merge_cells('AG9:AJ9')
        ws['AG9'] = telefone
        ws['AG9'].font = self.normal_font

        # Trimester label
        ws['AK9'] = "Trimestre:"
        ws['AK9'].font = self.subtitle_font

        # Parse period to extract trimester and year
        trimestre, ano = self._parse_period(self.period)
        ws.merge_cells('AM9:AP9')
        ws['AM9'] = trimestre
        ws['AM9'].font = self.normal_font

        # Year label
        ws['AQ9'] = "Ano:"
        ws['AQ9'].font = self.subtitle_font

        # Year value (merged cells)
        ws.merge_cells('AS9:AT9')
        ws['AS9'] = ano
        ws['AS9'].font = self.normal_font

    def _add_responsible_section(self, ws) -> None:
        """Add responsible person section (rows 11-12)"""
        # Row 11: Responsible person
        ws['A11'] = "Nome  e CPF do responsável pelo preenchimento:"
        ws['A11'].font = self.subtitle_font

        # This would come from user data
        responsible = self.company_info.get('responsavel', '')
        ws.merge_cells('K11:AT11')
        ws['K11'] = responsible
        ws['K11'].font = self.normal_font

        # Row 12: Location and date
        ws['A12'] = "Local e data:"
        ws['A12'].font = self.subtitle_font

        location_date = f"{self.company_info.get('cidade', 'CIDADE')}, {datetime.now().strftime('%d de %B de %Y')}"
        ws.merge_cells('D12:AT12')
        ws['D12'] = location_date
        ws['D12'].font = self.normal_font

    def _add_column_headers(self, ws) -> None:
        """
        Add multi-row column headers (rows 14-16).

        The MAPA format has a complex 3-row header structure:
        - Row 14: Main category headers (merged cells)
        - Row 15: Sub-category headers
        - Row 16: Specific field headers
        """
        # Row 14 - Main headers
        headers_row14 = [
            (1, 4, "MATERIAL/ PRODUTO"),
            (5, 7, "Nº REGISTRO DE PRODUTO/ AUTORIZAÇÃO"),
            (8, 24, "GARANTIAS (NUTRIENTES) - em %"),
            (25, 25, "Unidade"),
            (26, 26, "Estoque Inicial"),
            (30, 41, "QUANTIDADES DE MATÉRIAS-PRIMAS \nUso na fabricação de fertilizantes e Vendas como Produto Acabado"),
            (42, 43, "OUTRAS ENTRADAS"),
            (44, 45, "OUTRAS SAÍDAS"),
            (46, 46, "Estoque Final no Trimestre"),
        ]

        for start_col, end_col, text in headers_row14:
            if start_col == end_col:
                cell = ws.cell(row=14, column=start_col)
                cell.value = text
            else:
                ws.merge_cells(start_row=14, start_column=start_col, end_row=14, end_column=end_col)
                cell = ws.cell(row=14, column=start_col)
                cell.value = text

            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.fill = self.header_fill
            cell.border = self.thin_border

        # Row 15 - Sub headers
        nutrient_headers = [
            (8, "N"), (9, "P2O5"), (11, "K2O"), (12, "Ca"), (13, "Mg"),
            (14, "S"), (15, "B"), (16, "Cl"), (17, "Co"), (18, "Cu"),
            (19, "Fe"), (20, "Mn"), (21, "Mo"), (22, "Ni"), (23, "Si"), (24, "Zn")
        ]

        for col, text in nutrient_headers:
            cell = ws.cell(row=15, column=col)
            cell.value = text
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.fill = self.header_fill
            cell.border = self.thin_border

        # COMPRAS/ENTRADAS header
        ws.merge_cells(start_row=15, start_column=30, end_row=15, end_column=37)
        cell = ws.cell(row=15, column=30)
        cell.value = "COMPRAS/ENTRADAS"
        cell.font = self.header_font
        cell.alignment = self.center_alignment
        cell.fill = self.header_fill
        cell.border = self.thin_border

        # DESTINAÇÃO header
        ws.merge_cells(start_row=15, start_column=38, end_row=15, end_column=41)
        cell = ws.cell(row=15, column=38)
        cell.value = "DESTINAÇÃO"
        cell.font = self.header_font
        cell.alignment = self.center_alignment
        cell.fill = self.header_fill
        cell.border = self.thin_border

        # Row 16 - Specific headers
        specific_headers = [
            (9, "Total"), (10, "Solúvel"),
            (30, "NACIONAL"), (33, "NOME DA EMPRESA"), (34, "IMPORTADO"), (37, "País de origem"),
            (38, "Uso na Produção"), (41, "Venda como Produto"),
            (42, "Quantidade"), (43, "Discriminação"),
            (44, "Quantidade"), (45, "Discriminação"),
        ]

        for col, text in specific_headers:
            cell = ws.cell(row=16, column=col)
            cell.value = text
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.fill = self.header_fill
            cell.border = self.thin_border

    def _add_data_rows(self, ws) -> None:
        """Add product data rows (starting at row 17)"""
        current_row = 17

        for mapa_row in self.mapa_rows:
            # Column 1: Material/Produto (merged cells 1-4)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell = ws.cell(row=current_row, column=1)
            cell.value = mapa_row.material_produto
            cell.font = self.normal_font
            cell.alignment = self.left_alignment
            cell.border = self.thin_border

            # Column 5-7: Registro (merged)
            ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
            cell = ws.cell(row=current_row, column=5)
            cell.value = mapa_row.registro_produto
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Columns 8-24: Guarantees (Nutrients)
            guarantees = [
                mapa_row.n_total,           # 8: N
                mapa_row.p2o5_total,        # 9: P2O5 Total
                mapa_row.p2o5_soluvel,      # 10: P2O5 Solúvel
                mapa_row.k2o,               # 11: K2O
                mapa_row.ca,                # 12: Ca
                mapa_row.mg,                # 13: Mg
                mapa_row.s,                 # 14: S
                mapa_row.b,                 # 15: B
                mapa_row.cl,                # 16: Cl
                mapa_row.co,                # 17: Co
                mapa_row.cu,                # 18: Cu
                mapa_row.fe,                # 19: Fe
                mapa_row.mn,                # 20: Mn
                mapa_row.mo,                # 21: Mo
                mapa_row.ni,                # 22: Ni
                mapa_row.si,                # 23: Si
                mapa_row.zn,                # 24: Zn
            ]

            for idx, value in enumerate(guarantees, start=8):
                cell = ws.cell(row=current_row, column=idx)
                cell.value = value if value else ""
                cell.font = self.normal_font
                cell.alignment = self.center_alignment
                cell.border = self.thin_border

            # Column 25: Unidade
            cell = ws.cell(row=current_row, column=25)
            cell.value = mapa_row.unidade
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Column 26: Estoque Inicial
            cell = ws.cell(row=current_row, column=26)
            cell.value = mapa_row.estoque_inicial
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Columns 30-32: Nacional (merged)
            ws.merge_cells(start_row=current_row, start_column=30, end_row=current_row, end_column=32)
            cell = ws.cell(row=current_row, column=30)
            cell.value = float(mapa_row.compras_nacional) if mapa_row.compras_nacional else ""
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Column 33: Nome da Empresa
            cell = ws.cell(row=current_row, column=33)
            cell.value = mapa_row.compras_empresa_nome
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Columns 34-36: Importado (merged)
            ws.merge_cells(start_row=current_row, start_column=34, end_row=current_row, end_column=36)
            cell = ws.cell(row=current_row, column=34)
            cell.value = float(mapa_row.compras_importado) if mapa_row.compras_importado else ""
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Column 37: País de origem
            cell = ws.cell(row=current_row, column=37)
            cell.value = mapa_row.compras_pais_origem
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Columns 38-40: Uso na Produção (merged)
            ws.merge_cells(start_row=current_row, start_column=38, end_row=current_row, end_column=40)
            cell = ws.cell(row=current_row, column=38)
            cell.value = float(mapa_row.uso_na_producao) if mapa_row.uso_na_producao else ""
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Column 41: Venda como Produto
            cell = ws.cell(row=current_row, column=41)
            cell.value = float(mapa_row.venda_como_produto) if mapa_row.venda_como_produto else ""
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Column 42: Outras Entradas - Quantidade
            cell = ws.cell(row=current_row, column=42)
            cell.value = float(mapa_row.outras_entradas_quantidade) if mapa_row.outras_entradas_quantidade else ""
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Column 43: Outras Entradas - Discriminação
            cell = ws.cell(row=current_row, column=43)
            cell.value = mapa_row.outras_entradas_discriminacao
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Column 44: Outras Saídas - Quantidade
            cell = ws.cell(row=current_row, column=44)
            cell.value = float(mapa_row.outras_saidas_quantidade) if mapa_row.outras_saidas_quantidade else ""
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Column 45: Outras Saídas - Discriminação
            cell = ws.cell(row=current_row, column=45)
            cell.value = mapa_row.outras_saidas_discriminacao
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Column 46: Estoque Final
            cell = ws.cell(row=current_row, column=46)
            cell.value = mapa_row.estoque_final
            cell.font = self.normal_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            current_row += 1

    def _adjust_column_widths(self, ws) -> None:
        """Adjust column widths for readability"""
        # Material/Produto columns (1-4)
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 8

        # Registro columns (5-7)
        for col in range(5, 8):
            ws.column_dimensions[get_column_letter(col)].width = 7

        # Nutrient columns (8-24)
        for col in range(8, 25):
            ws.column_dimensions[get_column_letter(col)].width = 5

        # Unidade (25)
        ws.column_dimensions[get_column_letter(25)].width = 9

        # Other columns (26-46)
        for col in range(26, 47):
            ws.column_dimensions[get_column_letter(col)].width = 8

        # Company name column (33) - wider
        ws.column_dimensions[get_column_letter(33)].width = 15

        # Set row heights for headers
        ws.row_dimensions[14].height = 30
        ws.row_dimensions[15].height = 25
        ws.row_dimensions[16].height = 25

    def _parse_period(self, period: str) -> tuple:
        """
        Parse period string to extract trimester and year.

        Args:
            period: Period string like "1º-2025", "Q3-2025", "3º-2025"

        Returns:
            Tuple of (trimester, year)
        """
        # Handle different formats
        if 'Q' in period.upper():
            parts = period.upper().split('Q')
            if len(parts) == 2:
                year = parts[1].strip('-')
                quarter = parts[1][0] if len(parts[1]) > 0 else '1'
                return f"{quarter}º", year

        if 'º' in period:
            parts = period.split('º')
            trimester = parts[0].strip() + 'º'
            year = parts[1].strip('-') if len(parts) > 1 else datetime.now().year
            return trimester, str(year)

        # Default
        return "1º", str(datetime.now().year)

    def _save_workbook(self, wb: Workbook) -> str:
        """
        Save workbook to file.

        Args:
            wb: Workbook to save

        Returns:
            Path to saved file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Relatorio_MAPA_{self.period}_{timestamp}.xlsx"
        filepath = os.path.join(self.report_dir, f"user_{self.user.id}_{filename}")

        wb.save(filepath)

        return filepath

    def get_row_count(self) -> int:
        """Get number of data rows in report"""
        return len(self.mapa_rows)
